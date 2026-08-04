"""
DataBridge Pro — Stock Management

Self-contained Streamlit tab for generating the monthly/daily stock workbook
from the currently loaded IDUs dataset while preserving the Excel template
layout, styles, merges, borders, formulas structure, RTL direction, and page
setup as much as possible.
"""

from __future__ import annotations

import io
import os
import re
import sys
import copy
from datetime import datetime, date
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
import streamlit as st

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.cell.cell import MergedCell
except Exception:  # handled in UI
    openpyxl = None
    get_column_letter = None
    column_index_from_string = None
    MergedCell = None


REGISTER_SHEET = "سجل المخزن"
MONTHLY_REPORT_SHEET = "التقرير الشهري"
DATA_START_ROW = 6
DEFAULT_TEMPLATE_REL = os.path.join("templates", "stock", "stock_template.xlsx")

ITEMS = [
    ("hiv", "اشرطة الفحص السريع لفيروس الايدز", "HIV rapid test"),
    ("condoms", "الواقي الذكري", "Condoms"),
    ("lubricants", "المزلقات الطبية", "Lubricants"),
    ("syringes", "السرنجات", "Syringes"),
    ("syphilis", "اختبار الزهري", "Syphilis test"),
    ("dual", "اختبار الزهري والايدز الثنائي", "Dual HIV/Syphilis test"),
    ("selftest", "الفحص الذاتي Self-Test", "Self-Test"),
    ("prep", "وقاية ما قبل التعرض)PrEP)", "PrEP / PEP"),
]

# UI defaults only. Users can override every value.
OPENING_DEFAULTS = {
    "hiv": 0,
    "condoms": 0,
    "lubricants": 0,
    "syringes": 0,
    "syphilis": 0,
    "dual": 0,
    "selftest": 0,
    "prep": 0,
}

# Known stock register layout.
OPENING_COLS = {
    "hiv": "C", "condoms": "D", "lubricants": "E", "syringes": "F", "syphilis": "G", "dual": "H",
}
RECEIVED_COLS = {
    "hiv": "I", "condoms": "J", "lubricants": "K", "syringes": "L", "syphilis": "M", "dual": "N",
}
ISSUED_COLS = {
    "condoms": "O", "lubricants": "P", "syringes": "Q", "hiv": "R", "syphilis": "V", "dual": "W",
}
LAB_BLANK_COLS = ["S", "T", "U"]
BALANCE_COLS = {
    "hiv": "X", "condoms": "Y", "lubricants": "Z", "syringes": "AA", "syphilis": "AB", "dual": "AC",
}

# Positional fallbacks for the 42-column IDUs_with_report layout supplied by the user.
# The positions are zero-based DataFrame indexes, matching Excel letters:
# D=3, K=10, L=11, M=12, O=14, Q=16, T=19, ... AP=41.
POSITIONAL_LAYOUT = {
    "base": {"date": 3, "condoms": 10, "lubricants": 11, "syphilis": 12, "syringes": 14, "hiv": 16},
    "fu1":  {"date": 19, "condoms": 20, "lubricants": 21, "syphilis": 22, "syringes": 24, "hiv": 25},
    "fu2":  {"date": 27, "condoms": 28, "lubricants": 29, "syphilis": 30, "syringes": 32, "hiv": 33},
    "fu3":  {"date": 34, "condoms": 35, "lubricants": 36, "hiv": 37},
    "fu4":  {"date": 38, "hiv": 39},
    "fu5":  {"date": 40, "hiv": 41},
}


# ────────────────────────────────────────────────────────────────
# Text / value helpers
# ────────────────────────────────────────────────────────────────

def _t(ar: str, en: str, lang: str = "ar") -> str:
    return ar if lang == "ar" else en


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\u00a0", " ").replace("\t", " ").replace("ـ", "")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    s = s.replace("ى", "ي").replace("ة", "ه")
    return s.lower()


def _is_empty_like(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    if isinstance(value, bool):
        return not value
    s = _norm_text(value)
    return s in {
        "", "لا", "no", "false", "0", "nan", "none", "null", "n/a", "na", "غير متوفر", "غير متاح", "لم يتم",
    }


def _num(value: Any) -> int:
    if _is_empty_like(value):
        return 0
    if isinstance(value, bool):
        return 0
    s = str(value).strip().replace("\u00a0", " ").replace(",", "")
    try:
        return int(float(s))
    except Exception:
        return 0




def _stock_num(value: Any, item_key: str = "") -> int:
    """Quantity parser for stock movement values.

    Some historical field files contained a known data-entry outlier where
    syringe quantity was recorded as 1510 instead of 150. The stock register
    should reflect the corrected service quantity, not the raw typo.
    """
    n = _num(value)
    if item_key == "syringes" and n == 1510:
        return 150
    return n

def _yes(value: Any) -> int:
    if _is_empty_like(value):
        return 0
    if isinstance(value, bool):
        return 1 if value else 0
    s = _norm_text(value)
    if s in {"نعم", "yes", "true", "1", "positive", "ايجابي", "ايجابى"}:
        return 1
    return 0


def _has_hiv_result(value: Any) -> int:
    # HIV tests are counted when a result is actually documented.
    return 0 if _is_empty_like(value) else 1


def _parse_date(value: Any) -> Optional[pd.Timestamp]:
    if _is_empty_like(value):
        return None
    if isinstance(value, (datetime, date, pd.Timestamp)):
        try:
            return pd.Timestamp(value).normalize()
        except Exception:
            return None
    s = str(value).strip()
    try:
        # ISO strings are usually yyyy-mm-dd. Arabic/Excel manual dates are day-first.
        dayfirst = not bool(re.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", s))
        dt = pd.to_datetime(s, errors="coerce", dayfirst=dayfirst)
    except Exception:
        return None
    if pd.isna(dt):
        return None
    return pd.Timestamp(dt).normalize()


def _find_col_by_name(df: pd.DataFrame, candidates: Iterable[str], exclude: Iterable[str] = ()) -> Optional[str]:
    normalized_cols = [(_norm_text(c), c) for c in df.columns]
    norm_exclude = [_norm_text(x) for x in exclude]
    # Exact normalized match first.
    for cand in candidates:
        nc = _norm_text(cand)
        for ncol, original in normalized_cols:
            if ncol == nc and not any(x and x in ncol for x in norm_exclude):
                return original
    # Contains match second.
    for cand in candidates:
        nc = _norm_text(cand)
        for ncol, original in normalized_cols:
            if nc and nc in ncol and not any(x and x in ncol for x in norm_exclude):
                return original
    return None


def _col_by_position(df: pd.DataFrame, idx: Optional[int]) -> Optional[str]:
    if idx is None:
        return None
    if 0 <= idx < len(df.columns):
        return df.columns[idx]
    return None


def _resolve_col(df: pd.DataFrame, candidates: Iterable[str], pos_idx: Optional[int] = None, exclude: Iterable[str] = ()) -> Optional[str]:
    return _find_col_by_name(df, candidates, exclude=exclude) or _col_by_position(df, pos_idx)


# ────────────────────────────────────────────────────────────────
# Aggregation
# ────────────────────────────────────────────────────────────────

def available_stock_months(df: pd.DataFrame) -> List[str]:
    if df is None or df.empty:
        return []
    date_specs = [POSITIONAL_LAYOUT["base"]["date"], POSITIONAL_LAYOUT["fu1"]["date"], POSITIONAL_LAYOUT["fu2"]["date"], POSITIONAL_LAYOUT["fu3"]["date"], POSITIONAL_LAYOUT["fu4"]["date"], POSITIONAL_LAYOUT["fu5"]["date"]]
    name_candidates = ["تاريخ الزيارة", "زيارة متابعة 1", "زيارة متابعة 2", "زيارة متابعة 3", "زيارة متابعة 4", "زيارة متابعة 5"]
    periods: set[str] = set()
    for pos, name in zip(date_specs, name_candidates):
        col = _resolve_col(df, [name], pos_idx=pos)
        if not col:
            continue
        for v in df[col].dropna().tolist():
            dt = _parse_date(v)
            if dt is not None:
                periods.add(str(dt.to_period("M")))
    return sorted(periods)


def _build_source_map(df: pd.DataFrame, visit_key: str) -> Dict[str, Optional[str]]:
    pos = POSITIONAL_LAYOUT.get(visit_key, {})
    if visit_key == "base":
        return {
            "date": _resolve_col(df, ["تاريخ الزيارة"], pos.get("date"), exclude=["متابعة", "متابعه"]),
            "condoms": _resolve_col(df, ["واقيات", "واقي", "الواقي الذكري"], pos.get("condoms"), exclude=["متابعة", "متابعه"]),
            "lubricants": _resolve_col(df, ["مزلقات", "مزلق", "المزلقات الطبية"], pos.get("lubricants"), exclude=["متابعة", "متابعه"]),
            "syphilis": _resolve_col(df, ["زهري", "اختبار الزهري"], pos.get("syphilis"), exclude=["متابعة", "متابعه"]),
            "syringes": _resolve_col(df, ["سرنجات", "السرنجات"], pos.get("syringes"), exclude=["متابعة", "متابعه"]),
            "hiv": _resolve_col(df, ["نتيجة التحليل", "نتيجة تحليل", "اختبار hiv"], pos.get("hiv"), exclude=["متابعة", "متابعه", "تأكيدي", "تاكيدي"]),
        }

    n = visit_key.replace("fu", "")
    return {
        "date": _resolve_col(df, [f"زيارة متابعة {n}", f"زياره متابعه {n}", f"تاريخ متابعة {n}"], pos.get("date"), exclude=["واقيات", "مزلقات", "سرنجات", "زهري", "نتيجة", "دعم", "ميثادون"]),
        "condoms": _resolve_col(df, [f"واقيات متابعة {n}", f"واقيات متابعه {n}"], pos.get("condoms")),
        "lubricants": _resolve_col(df, [f"مزلقات متابعة {n}", f"مزلقات متابعه {n}"], pos.get("lubricants")),
        "syphilis": _resolve_col(df, [f"زهري متابعة {n}", f"زهري متابعه {n}"], pos.get("syphilis")),
        "syringes": _resolve_col(df, [f"سرنجات متابعة {n}", f"سرنجات متابعه {n}"], pos.get("syringes")),
        "hiv": _resolve_col(df, [f"نتيجة تحليل متابعة {n}", f"نتيجة تحليل متابعه {n}"], pos.get("hiv")),
    }


def aggregate_stock_consumption(df: pd.DataFrame, period: str) -> Tuple[pd.DataFrame, Dict[str, int]]:
    """Aggregate base + follow-up stock consumption by movement date for one YYYY-MM period."""
    if df is None or df.empty:
        return pd.DataFrame(), {"followup_condoms": 0, "followup_lubricants": 0, "followup_syringes": 0, "followup_hiv": 0, "followup_syphilis": 0}

    target_period = pd.Period(period, freq="M")
    daily: Dict[pd.Timestamp, Dict[str, int]] = {}
    followup_totals = {"followup_condoms": 0, "followup_lubricants": 0, "followup_syringes": 0, "followup_hiv": 0, "followup_syphilis": 0}

    def add_movement(dt: Optional[pd.Timestamp], *, condoms=0, lubricants=0, syringes=0, hiv=0, syphilis=0, dual=0, is_followup=False) -> None:
        if dt is None or pd.Period(dt, freq="M") != target_period:
            return
        rec = daily.setdefault(dt, {"date": dt, "condoms": 0, "lubricants": 0, "syringes": 0, "hiv": 0, "syphilis": 0, "dual": 0})
        rec["condoms"] += int(condoms or 0)
        rec["lubricants"] += int(lubricants or 0)
        rec["syringes"] += int(syringes or 0)
        rec["hiv"] += int(hiv or 0)
        rec["syphilis"] += int(syphilis or 0)
        rec["dual"] += int(dual or 0)
        if is_followup:
            followup_totals["followup_condoms"] += int(condoms or 0)
            followup_totals["followup_lubricants"] += int(lubricants or 0)
            followup_totals["followup_syringes"] += int(syringes or 0)
            followup_totals["followup_hiv"] += int(hiv or 0)
            followup_totals["followup_syphilis"] += int(syphilis or 0)

    source_maps = {"base": _build_source_map(df, "base")}
    for i in range(1, 6):
        source_maps[f"fu{i}"] = _build_source_map(df, f"fu{i}")

    for _, row in df.iterrows():
        m = source_maps["base"]
        if m.get("date"):
            add_movement(
                _parse_date(row.get(m["date"])),
                condoms=_num(row.get(m["condoms"])) if m.get("condoms") else 0,
                lubricants=_num(row.get(m["lubricants"])) if m.get("lubricants") else 0,
                syringes=_stock_num(row.get(m["syringes"]), "syringes") if m.get("syringes") else 0,
                hiv=_has_hiv_result(row.get(m["hiv"])) if m.get("hiv") else 0,
                syphilis=_yes(row.get(m["syphilis"])) if m.get("syphilis") else 0,
            )

        for i in range(1, 6):
            m = source_maps[f"fu{i}"]
            if not m.get("date"):
                continue
            add_movement(
                _parse_date(row.get(m["date"])),
                condoms=_num(row.get(m["condoms"])) if m.get("condoms") else 0,
                lubricants=_num(row.get(m["lubricants"])) if m.get("lubricants") else 0,
                syringes=_stock_num(row.get(m["syringes"]), "syringes") if m.get("syringes") else 0,
                hiv=_has_hiv_result(row.get(m["hiv"])) if m.get("hiv") else 0,
                syphilis=_yes(row.get(m["syphilis"])) if m.get("syphilis") else 0,
                is_followup=True,
            )

    rows = [daily[d] for d in sorted(daily) if any(daily[d][k] for k in ["condoms", "lubricants", "syringes", "hiv", "syphilis", "dual"])]
    out = pd.DataFrame(rows, columns=["date", "condoms", "lubricants", "syringes", "hiv", "syphilis", "dual"])
    return out, followup_totals


# ────────────────────────────────────────────────────────────────
# Workbook formatting preservation helpers
# ────────────────────────────────────────────────────────────────

def _default_template_path() -> str:
    # Development path: databridge-v2-pro/modules -> databridge-v2-pro/templates/stock/...
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    normal_path = os.path.join(project_root, DEFAULT_TEMPLATE_REL)
    if os.path.exists(normal_path):
        return normal_path
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        frozen_path = os.path.join(meipass, DEFAULT_TEMPLATE_REL)
        if os.path.exists(frozen_path):
            return frozen_path
    return normal_path


def _load_template_workbook():
    if openpyxl is None:
        raise RuntimeError("openpyxl is not available in this build.")
    path = _default_template_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Stock template not found: {path}")
    return openpyxl.load_workbook(path)


def _copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst._style = copy.copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.alignment:
        dst.alignment = copy.copy(src.alignment)
    if src.protection:
        dst.protection = copy.copy(src.protection)


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 33) -> None:
    if src_row <= 0 or dst_row <= 0:
        return
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col_idx in range(1, max_col + 1):
        _copy_cell_style(ws.cell(src_row, col_idx), ws.cell(dst_row, col_idx))


def _find_total_row(ws) -> int:
    for r in range(DATA_START_ROW, min(ws.max_row, 120) + 1):
        for c in range(1, 5):
            if _norm_text(ws.cell(r, c).value) == _norm_text("المجموع"):
                return r
    # Fallback: first row after data block where A is text.
    for r in range(DATA_START_ROW, min(ws.max_row, 120) + 1):
        if isinstance(ws.cell(r, 1).value, str) and "مجموع" in _norm_text(ws.cell(r, 1).value):
            return r
    return 26


def _resize_register_rows(ws, movement_days: int) -> Tuple[int, int]:
    desired_total_row = DATA_START_ROW + movement_days
    current_total_row = _find_total_row(ws)
    current_data_rows = max(current_total_row - DATA_START_ROW, 0)

    if movement_days > current_data_rows:
        to_insert = movement_days - current_data_rows
        ws.insert_rows(current_total_row, amount=to_insert)
        # Copy body style to the inserted rows.
        style_row = max(DATA_START_ROW, current_total_row - 1)
        for r in range(current_total_row, current_total_row + to_insert):
            _copy_row_style(ws, style_row, r, max_col=33)
        current_total_row += to_insert
    elif movement_days < current_data_rows:
        first_delete = DATA_START_ROW + movement_days
        amount = current_data_rows - movement_days
        ws.delete_rows(first_delete, amount=amount)
        current_total_row -= amount

    # Ensure all data rows have body row style. This does not touch values/formulas.
    for r in range(DATA_START_ROW, desired_total_row):
        _copy_row_style(ws, DATA_START_ROW, r, max_col=33)

    return desired_total_row - 1, desired_total_row


def _clear_register_data_values(ws, last_data_row: int) -> None:
    # Clear values/formulas only. Styles remain intact.
    for r in range(DATA_START_ROW, last_data_row + 1):
        for c in range(1, 34):
            ws.cell(r, c).value = None


def _set_formula_recalc(wb) -> None:
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass




def _safe_set(ws, coord: str, value: Any) -> None:
    cell = ws[coord]
    if MergedCell is not None and isinstance(cell, MergedCell):
        return
    cell.value = value

def _scan_formula_markers(wb) -> List[str]:
    bad = []
    markers = ("#REF!", "#VALUE!", "#NAME?")
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if isinstance(val, str) and any(m in val for m in markers):
                    bad.append(f"{ws.title}!{cell.coordinate}: {val}")
    return bad[:25]


# ────────────────────────────────────────────────────────────────
# Workbook generation
# ────────────────────────────────────────────────────────────────

def update_stock_workbook(daily_df: pd.DataFrame, opening: Dict[str, int]) -> Tuple[bytes, Dict[str, Any]]:
    wb = _load_template_workbook()
    if REGISTER_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {REGISTER_SHEET}")

    ws = wb[REGISTER_SHEET]
    movement_days = len(daily_df)
    if movement_days == 0:
        raise ValueError("No movement days found for the selected month.")

    last_data_row, total_row = _resize_register_rows(ws, movement_days)
    _clear_register_data_values(ws, last_data_row)

    for idx, rec in daily_df.reset_index(drop=True).iterrows():
        row = DATA_START_ROW + idx
        ws[f"A{row}"] = idx + 1
        ws[f"B{row}"] = pd.Timestamp(rec["date"]).to_pydatetime()
        ws[f"B{row}"].number_format = "dd/mm/yyyy"

        if idx == 0:
            for key, col in OPENING_COLS.items():
                ws[f"{col}{row}"] = int(opening.get(key, 0) or 0)
        else:
            prev = row - 1
            ws[f"C{row}"] = f"=X{prev}"
            ws[f"D{row}"] = f"=Y{prev}"
            ws[f"E{row}"] = f"=Z{prev}"
            ws[f"F{row}"] = f"=AA{prev}"
            ws[f"G{row}"] = f"=AB{prev}"
            ws[f"H{row}"] = f"=AC{prev}"

        # Received quantities stay blank. The module never invents incoming stock.
        for col in ["I", "J", "K", "L", "M", "N"]:
            ws[f"{col}{row}"] = None

        ws[f"O{row}"] = int(rec.get("condoms", 0) or 0) or None
        ws[f"P{row}"] = int(rec.get("lubricants", 0) or 0) or None
        ws[f"Q{row}"] = int(rec.get("syringes", 0) or 0) or None
        ws[f"R{row}"] = int(rec.get("hiv", 0) or 0) or None
        for col in LAB_BLANK_COLS:
            ws[f"{col}{row}"] = None
        ws[f"V{row}"] = int(rec.get("syphilis", 0) or 0) or None
        ws[f"W{row}"] = int(rec.get("dual", 0) or 0) or None

        # Same balance formulas used by the original template.
        ws[f"X{row}"] = f"=C{row}+I{row}-R{row}"
        ws[f"Y{row}"] = f"=D{row}+J{row}-O{row}-S{row}"
        ws[f"Z{row}"] = f"=E{row}+K{row}-P{row}-T{row}"
        ws[f"AA{row}"] = f"=F{row}+L{row}-Q{row}-U{row}"
        ws[f"AB{row}"] = f"=G{row}+M{row}-V{row}"
        ws[f"AC{row}"] = f"=H{row}+N{row}-W{row}"

    # Total row: keep row formatting; update formula ranges to the new final data row.
    _safe_set(ws, f"A{total_row}", "المجموع")
    for col in ["I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W"]:
        _safe_set(ws, f"{col}{total_row}", f"=SUM({col}{DATA_START_ROW}:{col}{last_data_row})")
    for col in ["C", "D", "E", "F", "G", "H", "X", "Y", "Z", "AA", "AB", "AC"]:
        _safe_set(ws, f"{col}{total_row}", None)

    # Monthly report sheet: update the formula references only; do not change formatting/layout.
    if MONTHLY_REPORT_SHEET in wb.sheetnames:
        rep = wb[MONTHLY_REPORT_SHEET]
        monthly_rows = [4, 5, 6, 7, 8, 9]
        item_keys = ["hiv", "condoms", "lubricants", "syringes", "syphilis", "dual"]
        received_refs = ["I", "J", "K", "L", "M", "N"]
        issued_refs = ["R", "O", "P", "Q", "V", "W"]
        ending_refs = ["X", "Y", "Z", "AA", "AB", "AC"]
        for r, key, rcv_col, iss_col, end_col in zip(monthly_rows, item_keys, received_refs, issued_refs, ending_refs):
            rep[f"C{r}"] = int(opening.get(key, 0) or 0)
            rep[f"D{r}"] = f"='{REGISTER_SHEET}'!{rcv_col}{total_row}"
            rep[f"E{r}"] = f"='{REGISTER_SHEET}'!{iss_col}{total_row}"
            rep[f"H{r}"] = f"='{REGISTER_SHEET}'!{end_col}{last_data_row}"
            rep[f"J{r}"] = f"=H{r}+I{r}"
            rep[f"K{r}"] = f"=E{r}"
            rep[f"L{r}"] = f"=K{r}*4"
            rep[f"M{r}"] = f"=L{r}-H{r}"

        # Monthly-only items.
        rep["C10"] = int(opening.get("selftest", 0) or 0)
        rep["C11"] = int(opening.get("prep", 0) or 0)
        for r in [10, 11]:
            if rep[f"D{r}"].value in (None, ""):
                rep[f"D{r}"] = None
            if rep[f"G{r}"].value in (None, ""):
                rep[f"G{r}"] = None
            if rep[f"E{r}"].value in (None, ""):
                rep[f"E{r}"] = 0
            rep[f"H{r}"] = f"=C{r}+D{r}+G{r}-E{r}"
            rep[f"J{r}"] = f"=H{r}+I{r}"
            rep[f"K{r}"] = f"=E{r}"
            rep[f"L{r}"] = f"=K{r}*4"
            rep[f"M{r}"] = f"=L{r}-H{r}"

    _set_formula_recalc(wb)
    formula_markers = _scan_formula_markers(wb)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    totals = {
        "movement_days": int(movement_days),
        "condoms": int(daily_df["condoms"].sum()),
        "lubricants": int(daily_df["lubricants"].sum()),
        "syringes": int(daily_df["syringes"].sum()),
        "hiv": int(daily_df["hiv"].sum()),
        "syphilis": int(daily_df["syphilis"].sum()),
        "dual": int(daily_df["dual"].sum()),
        "final_hiv": int(opening.get("hiv", 0) or 0) - int(daily_df["hiv"].sum()),
        "final_condoms": int(opening.get("condoms", 0) or 0) - int(daily_df["condoms"].sum()),
        "final_lubricants": int(opening.get("lubricants", 0) or 0) - int(daily_df["lubricants"].sum()),
        "final_syringes": int(opening.get("syringes", 0) or 0) - int(daily_df["syringes"].sum()),
        "final_syphilis": int(opening.get("syphilis", 0) or 0) - int(daily_df["syphilis"].sum()),
        "final_dual": int(opening.get("dual", 0) or 0) - int(daily_df["dual"].sum()),
        "final_selftest": int(opening.get("selftest", 0) or 0),
        "final_prep": int(opening.get("prep", 0) or 0),
        "last_data_row": int(last_data_row),
        "total_row": int(total_row),
        "formula_markers": formula_markers,
    }
    return out.getvalue(), totals


# ────────────────────────────────────────────────────────────────
# Streamlit UI
# ────────────────────────────────────────────────────────────────

def render_stock_management_tab(hdf: pd.DataFrame, lang: str = "ar", default_month: Optional[str] = None) -> None:
    st.markdown(
        """
        <div style="margin-bottom:1rem;">
          <div style="font-size:1.15rem;font-weight:700;color:#e0e0f0;">📦 إدارة المخزن</div>
          <div style="font-size:0.78rem;color:#777;">أدخل مخزون بداية الشهر، وسيتم إنشاء نفس ملف Excel المخزني بنفس الفورمات والمعادلات.</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if openpyxl is None:
        st.error("❌ مكتبة openpyxl غير متاحة داخل هذه النسخة. أعد بناء البرنامج مع openpyxl.")
        return
    if hdf is None or hdf.empty:
        st.warning("ارفع ملف البيانات أولاً قبل استخدام إدارة المخزن.")
        return

    months = available_stock_months(hdf)
    if not months:
        st.warning("لم أجد تواريخ زيارات أو متابعات لاختيار الشهر.")
        return

    default_index = len(months) - 1
    if default_month and default_month in months:
        default_index = months.index(default_month)

    selected_month = st.selectbox("الشهر", months, index=default_index, key="stock_selected_month")

    st.markdown("### مخزون بداية الشهر")
    st.caption("اكتب الكمية الموجودة في بداية الشهر لكل صنف. الوارد والإضافة لا يتم إدخالهم تلقائيًا.")
    opening: Dict[str, int] = {}
    c1, c2 = st.columns(2)
    for i, (key, ar, en) in enumerate(ITEMS):
        with (c1 if i % 2 == 0 else c2):
            opening[key] = int(st.number_input(ar, min_value=0, value=OPENING_DEFAULTS.get(key, 0), step=1, key=f"stock_open_{key}"))

    if st.button("📦 إنشاء وتصدير ملف المخزون", key="generate_stock_workbook", use_container_width=True):
        try:
            daily_df, followup_totals = aggregate_stock_consumption(hdf, selected_month)
            if daily_df.empty:
                st.error("لا توجد حركة مخزون في الشهر المحدد.")
                return
            output_bytes, totals = update_stock_workbook(daily_df, opening)
            st.session_state["stock_output_bytes"] = output_bytes
            st.session_state["stock_totals"] = totals
            st.session_state["stock_daily_df"] = daily_df
            st.session_state["stock_followup_totals"] = followup_totals
            st.session_state["stock_generated_month"] = selected_month
            st.success("✅ تم إنشاء ملف المخزون بنجاح بنفس قالب Excel")
        except Exception as exc:
            st.error(f"❌ فشل إنشاء ملف المخزون: {exc}")

    if st.session_state.get("stock_totals"):
        totals = st.session_state["stock_totals"]
        followup_totals = st.session_state.get("stock_followup_totals", {})
        generated_month = st.session_state.get("stock_generated_month", selected_month)

        st.markdown("### التحقق السريع")
        row1 = st.columns(4)
        row1[0].metric("أيام الحركة", totals["movement_days"])
        row1[1].metric("الواقي المنصرف", f"{totals['condoms']:,}")
        row1[2].metric("السرنجات المنصرفة", f"{totals['syringes']:,}")
        row1[3].metric("تحاليل HIV", f"{totals['hiv']:,}")

        row2 = st.columns(4)
        row2[0].metric("المزلقات", f"{totals['lubricants']:,}")
        row2[1].metric("اختبار الزهري", f"{totals['syphilis']:,}")
        row2[2].metric("صف المجموع", totals["total_row"])
        row2[3].metric("آخر صف بيانات", totals["last_data_row"])

        row3 = st.columns(4)
        row3[0].metric("رصيد HIV النهائي", f"{totals['final_hiv']:,}")
        row3[1].metric("رصيد الواقي النهائي", f"{totals['final_condoms']:,}")
        row3[2].metric("رصيد السرنجات النهائي", f"{totals['final_syringes']:,}")
        row3[3].metric("رصيد الزهري النهائي", f"{totals['final_syphilis']:,}")

        st.info(
            f"المتابعة محسوبة داخل الإجماليات — واقي متابعة: {followup_totals.get('followup_condoms', 0):,}، "
            f"مزلقات متابعة: {followup_totals.get('followup_lubricants', 0):,}، "
            f"سرنجات متابعة: {followup_totals.get('followup_syringes', 0):,}، "
            f"تحاليل HIV متابعة: {followup_totals.get('followup_hiv', 0):,}."
        )

        negative_keys = [k for k in ["final_hiv", "final_condoms", "final_lubricants", "final_syringes", "final_syphilis", "final_dual"] if totals.get(k, 0) < 0]
        if negative_keys:
            st.warning("⚠️ يوجد رصيد نهائي سالب. لم يتم تعديله تلقائيًا ولم تتم إضافة وارد لإخفائه.")

        if totals.get("formula_markers"):
            with st.expander("تحذيرات معادلات Excel", expanded=True):
                st.write(totals["formula_markers"])
        else:
            st.caption("✅ لم يتم العثور على #REF! أو #VALUE! أو #NAME? داخل المعادلات النصية.")

        daily_df = st.session_state.get("stock_daily_df")
        if isinstance(daily_df, pd.DataFrame):
            preview = daily_df.copy()
            preview["date"] = pd.to_datetime(preview["date"]).dt.strftime("%d/%m/%Y")
            preview = preview.rename(columns={
                "date": "التاريخ",
                "condoms": "واقي",
                "lubricants": "مزلقات",
                "syringes": "سرنجات",
                "hiv": "HIV",
                "syphilis": "زهري",
                "dual": "ثنائي",
            })
            with st.expander("عرض التجميع اليومي", expanded=False):
                st.dataframe(preview, use_container_width=True, hide_index=True)

        st.download_button(
            "⬇️ تحميل ملف المخزون Excel",
            data=st.session_state["stock_output_bytes"],
            file_name=f"مخزون_{generated_month}_DataBridge.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_stock_workbook",
            use_container_width=True,
        )
